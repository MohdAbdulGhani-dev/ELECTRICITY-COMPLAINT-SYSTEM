from flask import Flask , render_template,request,redirect,send_from_directory
from openpyxl import Workbook,load_workbook

import os

app=Flask(__name__)
excel_file="Registration_data.xlsx"

if not os.path.exists(excel_file): 
    wb= Workbook()
    ws=wb.active
    ws.append(["FULL NAME","USER NAME","EMAIL ID"
               ,"PHONE NUMBER","ADDRESS","PASSWORD","CONFIRM PASSWORD","ROLE"])
    wb.save(excel_file)



@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static','favicon.ico')

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/%20register')
def fix_register():
    return redirect('/register')



@app.route('/register',methods=['GET','POST'])
def register():
    if request.method=='POST':
        name=request.form.get('name')
        uname=request.form.get('uname')
        email=request.form.get('email')
        phone=request.form.get('phone')
        address=request.form.get('address')
        password=request.form.get('password')
        confirm_password=request.form.get('confirm_password')
        role=request.form.get('role')
        

        wb=load_workbook(excel_file)
        ws=wb.active
        ws.append([name,uname,email,phone,address,password,confirm_password,role])
        wb.save(excel_file)

        return render_template('success.html')
    return  render_template('register.html')







if __name__ =='__main__':
      app.run(debug=True)




